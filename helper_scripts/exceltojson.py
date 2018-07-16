from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='2017.xlsx')
print(wb.sheetnames)

sheet = wb.active

print(sheet['B2'].value)

data_list = []

for row in sheet.iter_rows(min_row=1, min_col=1, max_row=16386, max_col=13):
    a = []
    for cell in row:
        a.append(cell.value)
        # print cell.value
    # print type(a)
    post_data = {
            'last_name': a[1],
            'first_name': a[2],
            'middle_name': a[3],
            'dept': a[4],
            'employee_group': a[9],
            'compensation': a[10]
            }
    # print post_data
    data_list.append(post_data)
    
f = open("2017.txt", "w+");

f.write("{ \n\t\"data\": [\n");
for x in range(0, len(data_list)):
    person = data_list[x]
    # print person
    if person['first_name'] == None:
        person['first_name'] = ""
    if person['middle_name']== None:
        person['middle_name'] = ""
    if person['last_name'] == None:
        person['first_name'] = ""
    try:
        f.write("\t\t[\"" + person['last_name'] + "\",\"" + person['first_name'] + "\",\"" + person['middle_name'] + "\",\"" + person['employee_group'] + "\",\"" + person['dept'] + "\",\"" + str('${:,.2f}'.format(person['compensation'])) + "\"]")
        if(x < len(data_list)-1):
            f.write(",\n")
        else:
            f.write("\n")
    except Exception as e:
        print person
        print e
f.write("\t]\n}")
f.close()

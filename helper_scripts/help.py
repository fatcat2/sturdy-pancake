import json
import sqlite3
from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='2016.xlsx')

sheet = wb.active

data_list = []

#initiate connection with sqlite3 db file
conn = sqlite3.connect("salary_history.db")
c = conn.cursor()

#go through excel file and add entries to sqlite3 db
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=13292, max_col=5):
    a = []
    for cell in row:
        a.append(cell.value)
    
    post_data = {
            "last_name": a[2],
            "first_name": a[3],
            "middle_name": a[4],
            "dept": a[5],
            "employee_group": a[10],
            "compensation": a[12]
            }
    data_list.append(post_data)
    
f = open("ajax.txt", "w+");

f.write("{ \n\t\"data\": [\n");
for x in range(0, len(data_list)):
    person = data_list[x]
    # print person
    if person['first_name'] == None:
        person['first_name'] = ""
    if person['last_name'] == None:
        person['first_name'] = ""
    try:
        f.write("\t\t[\"" + person['last_name'] + "\",\"" + person['first_name'] + "\",\"" + person['employee_group'] + "\",\"" + person['dept'] + "\",\"" + str('${:,.2f}'.format(person['compensation'])) + "\"]")
        if(x < len(data_list)-1):
            f.write(",\n")
        else:
            f.write("\n")
    except Exception as e:
        print person
        print e
f.write("\t]\n}")
f.close()

import json
import sqlite3
from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='2017.xlsx')

sheet = wb.active

data_list = []

#initiate connection with sqlite3 db file
conn = sqlite3.connect("salary_history.db")
c = conn.cursor()

x = 200000000



#go through excel file and add entries to sqlite3 db
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=16695, max_col=13):
    a = []
    for cell in row:
        a.append(cell.value)
    compensation_json = {}
    compensation_json["2017"] = a[10]
    # print json.dumps(compensation_json)

    if a[3] == None:
        a[3] = ""
    if a[1] == None and a[2] == "None":
        continue
    try:
        post_data = {
                "id": x,
                "name": a[1]+a[2]+a[3],
                "last_name": a[1],
                "first_name": a[2],
                "middle_name": a[3],
                "dept": a[4],
                "employee_group": a[9],
                "compensation": json.dumps(compensation_json)
                }
    except Exception as e:
        print a
        print e
    data_list.append(post_data)
    x += 1

#add 2011 entries to sqlite3
for i in range(0, len(data_list)):
    person = data_list[i]
    try:
        insert_data = (person["id"], person["name"], person["last_name"], person["first_name"], person["middle_name"], person["dept"], person["employee_group"], person["compensation"])
        c.execute("insert into salary values(?, ?, ?, ?, ?, ?, ?, ?)", insert_data)
    except Exception as e:
        # print person
        print e
    conn.commit()

conn.close()

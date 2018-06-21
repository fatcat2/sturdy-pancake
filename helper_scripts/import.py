from pymongo.errors import BulkWriteError
from  pymongo import MongoClient
from openpyxl import Workbook
from openpyxl import load_workbook
import os
client = MongoClient('mongodb+srv://exponent_cand:'+os.environ.get('password')+'@cluster0-ufki0.mongodb.net/test')
# added a test comment
db = client["test"]
money = db.fourteen
wb = load_workbook(filename='2014.xlsx')
print(wb.sheetnames)

sheet = wb.active

print(sheet['B2'].value)

data_list = []

for row in sheet.iter_rows(min_row=2, min_col=3, max_row=16429, max_col=12):
    a = []
    for cell in row:
        a.append(cell.value)
    print type(a)
    post_data = {
            'last_name': a[0],
            'first_name': a[1],
            'middle_name': a[2],
            'dept': a[3],
            'city': a[5],
            'employee_group': a[8],
            'compensation': a[9]
            }
    data_list.append(post_data)
    #insert the entire fucking list yooooo
try:
    money.insert_many(data_list)
except BulkWriteError as bwe:
    print(bwe.details)

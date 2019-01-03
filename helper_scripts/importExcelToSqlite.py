import sqlite3
import json
from openpyxl import Workbook, load_workbook

# gather file info
year = input("Which year would you like to input?  ")
maxRows = int(input("How many rows are in the file?   "))
maxCols = int(input("How many columns are in the file?   "))
lastNameCol = int(input("What column is the last name in?   "))
firstNameCol = int(input("What column is the first name in?   "))
middleNameCol = int(input("What column is the middle name in?   "))
deptCol = int(input("What column is the department in?   "))
groupCol = int(input("What column is the group in?   "))
compCol = int(input("What column is the compensation in?   "))

# declare variables to be used
conn = sqlite3.connect("../static/salaries.db")
filename = "../excel/" + year + ".xlsx"
wb = load_workbook(filename=filename)
sheet = wb.active

# create le database
print("Creating database for year " + year)
conn.execute("create table Year"+year+" (lastName text, firstName text, middleName text, department text, empGroup text, compensation text)")

# populate le database
print("Populating database for year " + year)
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=maxRows, max_col=maxCols):
    a = []
    for cell in row:
        a.append(cell.value)
    print(a)
    print(maxRows, maxCols)
    if a[lastNameCol] == None:
        a[lastNameCol] = ""
    if a[firstNameCol] == None:
        a[firstNameCol] = ""
    if a[middleNameCol] == None:
        a[middleNameCol] == None

    conn.execute("insert into Year"+year+" (lastName, firstName, middleName, department, empGroup, compensation) values (?, ?, ?, ?, ?, ?)", (a[lastNameCol], a[firstNameCol], a[middleNameCol], a[deptCol], a[groupCol], a[compCol],))

conn.commit()
conn.close()

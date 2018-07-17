import json
import sqlite3
from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='2017.xlsx')

sheet = wb.active

data_list = []

#initiate connection with sqlite3 db file
conn = sqlite3.connect("salary_history.db")
c = conn.cursor()

x = 1500000

print "adding data"
#go through excel file and add entries to sqlite3 db
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=16695, max_col=13):
    a = []
    for cell in row:
        a.append(cell.value)
    compensation_json = {}
    compensation_json["2017"] = a[10]

    if a[3] == None:
        a[3] = ""

    if a[1] == None and a[2] == None:
        continue

    try:
        post_data = {
                "id": x,
                "name": a[1]+a[2]+a[3],
                "last_name": a[1],
                "first_name": a[2],
                "middle_name": a[3],
                "dept": a[4],
                "employee_group": a[9],
                "compensation_num": a[10],
                "compensation": json.dumps(compensation_json)
                }

        data_list.append(post_data)
    except Exception as e:
        print a
        print e
    x += 1

#add 2011 entries to sqlite3
for i in range(0, len(data_list)):
    person = data_list[i]
    try:
        insert_data = (person["id"], person["name"], person["last_name"], person["first_name"], person["middle_name"], person["dept"], person["employee_group"], person["compensation"])
        c.execute("select compensation from salary where name=?", (person["name"],))
        data = c.fetchone()
        # print data[0]
        if data:
            print person["name"] + " found in database"
            # print data[0]
            json_data = json.loads(data[0])
            # print type(json_data)
            json_data["2017"] = person["compensation_num"]
            print json_data["2017"], json_data
            c.execute("update salary set compensation=? where name=?", (str(json.dumps(json_data, sort_keys=True)), person["name"]))
        else:
            print "----------------"
            print person["name"] + " not found in db"
            print "----------------"
            c.execute("insert into salary values (?, ?, ?, ?, ?, ?, ?, ?)", (insert_data))
    except Exception as e:
        # print person
        print e

conn.commit()
conn.close()

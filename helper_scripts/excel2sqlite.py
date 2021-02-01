import sqlite3
import json
from openpyxl import Workbook, load_workbook

class Employee:
    def __init__(self, first_name, middle_name, last_name, department, group, compensation):
        self.first_name = first_name
        self.middle_name = middle_name
        self.last_name = last_name
        self.department = department
        self.group = group
        self.compensation = compensation
        self.long_text = f"{first_name} {last_name} in {department} in group {group} made ${compensation}"

# gather file info
year = input("Which year would you like to input?  ")
maxRows = int(input("How many rows are in the file?   "))
maxCols = int(input("How many columns are in the file?   "))
lastNameCol = int(input("What column is the last name in?   "))
firstNameCol = int(input("What column is the first name in?   "))
middleNameCol = int(input("What column is the middle name in?   "))
deptCol = int(input("What column is the department in?   "))
groupCol = int(input("What column is the group in?   "))
compCol = int(input("What column is the compensation in?   "))

# declare variables to be used
print("Connecting to database")
conn = sqlite3.connect("../data/salaries.db")
print("Connected to ../data/salaries.db")
print("Connecting to ../excel/"+year+".xlsx")
filename = "../excel/" + year + ".xlsx"
wb = load_workbook(filename=filename)
sheet = wb.active
print("Connected to excel sheet")

# create le database
print("Creating database for year " + year)
conn.execute(f"drop table if exists Year{year}")
conn.execute("create table Year"+year+" (lastName text, firstName text, middleName text, department text, empGroup text, compensation double, long_text text)")

# populate le database
print("Populating database for year " + year)
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=maxRows, max_col=maxCols):
    a = []
    for cell in row:
        a.append(cell.value)
    print(a)
    print(maxRows, maxCols)
    if a[lastNameCol] == None:
        a[lastNameCol] = ""
    if a[firstNameCol] == None:
        a[firstNameCol] = ""
    if a[middleNameCol] == None:
        a[middleNameCol] == None

    # make sure to skip students (only a 2016.xlsx problem)
    if a[groupCol] == "Student":
        continue

    emp = Employee(a[firstNameCol], a[middleNameCol], a[lastNameCol], a[deptCol], a[groupCol], a[compCol])

    conn.execute(
        "insert into Year"+year+" (lastName, firstName, middleName, department, empGroup, compensation, long_text) values (?, ?, ?, ?, ?, ?, ?)",
        (
            emp.last_name, emp.first_name, emp.middle_name, emp.department, emp.group, emp.compensation, emp.long_text,)
    )

conn.commit()
conn.close()
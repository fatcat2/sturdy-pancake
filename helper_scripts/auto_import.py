#!/usr/bin/env python3
"""
Automated Excel to SQLite importer for salary data.
Processes all Excel files in the excel/ directory with zero user interaction.
"""

import sqlite3
import os
import sys
from openpyxl import load_workbook

# Column header patterns to match (case-insensitive)
HEADER_PATTERNS = {
    'last_name': ['last name', 'last', 'lastname'],
    'first_name': ['first name', 'first', 'firstname'],
    'middle_name': ['mi', 'middle', 'middle name', 'middlename'],
    'department': ['department', 'dept'],
    'group': ['employee group', 'empgroup', 'title', 'group'],
    'compensation': ['compensation', 'comp'],
    'campus': ['campus'],
}


def find_column_index(headers, patterns):
    """Find the column index matching any of the given patterns."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]
    for i, header in enumerate(headers_lower):
        for pattern in patterns:
            if header == pattern:
                return i
    return None


def detect_columns(headers):
    """Auto-detect column positions from headers."""
    columns = {}
    for field, patterns in HEADER_PATTERNS.items():
        idx = find_column_index(headers, patterns)
        columns[field] = idx
    return columns


def parse_department(dept_value):
    """Parse department string to extract campus and department name."""
    if not dept_value:
        return '', ''

    dept_str = str(dept_value).strip()

    # Format: "Campus - Department" (e.g., "FW - English & Linguistic")
    if ' - ' in dept_str:
        parts = dept_str.split(' - ', 1)
        return parts[0].strip(), parts[1].strip()

    return '', dept_str


def process_excel_file(filepath, conn, year):
    """Process a single Excel file and import into database."""
    print(f"Processing {year}...")

    wb = load_workbook(filename=filepath, read_only=True)
    sheet = wb.active

    # Get headers from first row
    rows = list(sheet.iter_rows(min_row=1, max_row=1))
    if not rows:
        print(f"  Warning: Empty file for {year}")
        return 0

    headers = [cell.value for cell in rows[0]]
    columns = detect_columns(headers)

    # Validate required columns
    required = ['last_name', 'first_name', 'department', 'compensation']
    missing = [f for f in required if columns.get(f) is None]
    if missing:
        print(f"  Error: Missing required columns for {year}: {missing}")
        print(f"  Found headers: {headers}")
        return 0

    # Drop existing table and create new one
    conn.execute(f"DROP TABLE IF EXISTS Year{year}")
    conn.execute(f"""
        CREATE TABLE Year{year} (
            lastName TEXT,
            firstName TEXT,
            middleName TEXT,
            department TEXT,
            empGroup TEXT,
            compensation REAL,
            long_text TEXT,
            campus TEXT
        )
    """)

    # Process data rows
    row_count = 0
    for row in sheet.iter_rows(min_row=2):
        cells = [cell.value for cell in row]

        # Skip empty rows
        if not any(cells):
            continue

        # Extract values using detected column positions
        last_name = cells[columns['last_name']] or ''
        first_name = cells[columns['first_name']] or ''
        middle_name = cells[columns['middle_name']] if columns.get('middle_name') is not None else ''
        middle_name = middle_name or ''

        dept_raw = cells[columns['department']] or ''
        dept_campus, department = parse_department(dept_raw)

        group = ''
        if columns.get('group') is not None:
            group = cells[columns['group']] or ''

        # Skip students (2016 issue mentioned in original script)
        if group == 'Student':
            continue

        compensation = cells[columns['compensation']]
        if compensation is None:
            continue
        try:
            compensation = float(compensation)
        except (ValueError, TypeError):
            continue

        # Campus: prefer explicit column, fall back to parsed from department
        campus = ''
        if columns.get('campus') is not None:
            campus = cells[columns['campus']] or ''
        if not campus:
            campus = dept_campus

        long_text = f"{first_name} {last_name} in {department} in group {group} made ${compensation}"

        conn.execute(
            f"INSERT INTO Year{year} (lastName, firstName, middleName, department, empGroup, compensation, long_text, campus) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (last_name, first_name, middle_name, department, group, compensation, long_text, campus or 'None')
        )
        row_count += 1

    wb.close()
    print(f"  Imported {row_count} rows")
    return row_count


def create_auxiliary_tables(conn, year):
    """Create Department and Group lookup tables for a year."""
    conn.execute(f"DROP TABLE IF EXISTS Department{year}")
    conn.execute(f"DROP TABLE IF EXISTS Group{year}")

    conn.execute(f"CREATE TABLE Department{year} AS SELECT DISTINCT department FROM Year{year}")
    conn.execute(f"CREATE TABLE Group{year} AS SELECT DISTINCT empGroup FROM Year{year}")


def update_years_table(conn, years):
    """Update the years lookup table."""
    conn.execute("DROP TABLE IF EXISTS years")
    conn.execute("CREATE TABLE years (year INTEGER)")
    for year in sorted(years):
        conn.execute("INSERT INTO years (year) VALUES (?)", (year,))


def main():
    # Paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_dir = os.path.join(project_root, 'excel')
    db_path = os.path.join(project_root, 'data', 'salaries.db')

    # Parse command line args for specific years
    specific_years = None
    if len(sys.argv) > 1:
        try:
            specific_years = [int(y) for y in sys.argv[1:]]
            print(f"Processing specific years: {specific_years}")
        except ValueError:
            print("Usage: python auto_import.py [year1] [year2] ...")
            print("  No arguments = process all Excel files")
            sys.exit(1)

    # Find all Excel files
    excel_files = {}
    for filename in os.listdir(excel_dir):
        if filename.endswith('.xlsx'):
            try:
                year = int(filename.replace('.xlsx', ''))
                if specific_years is None or year in specific_years:
                    excel_files[year] = os.path.join(excel_dir, filename)
            except ValueError:
                continue

    if not excel_files:
        print("No Excel files found to process")
        sys.exit(1)

    # If no specific years requested, only process the newest year
    if specific_years is None:
        newest_year = max(excel_files.keys())
        excel_files = {newest_year: excel_files[newest_year]}
        print(f"Processing newest year only: {newest_year}")
    else:
        print(f"Found {len(excel_files)} Excel file(s) to process")
    print(f"Database: {db_path}")
    print()

    # Connect to database
    conn = sqlite3.connect(db_path)

    total_rows = 0
    processed_years = []

    for year in sorted(excel_files.keys()):
        filepath = excel_files[year]
        rows = process_excel_file(filepath, conn, year)
        if rows > 0:
            create_auxiliary_tables(conn, year)
            processed_years.append(year)
            total_rows += rows

    # Update years table with all years in database
    conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Year%'")
    all_years = []
    for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Year%'"):
        try:
            all_years.append(int(row[0].replace('Year', '')))
        except ValueError:
            continue

    update_years_table(conn, all_years)

    conn.commit()
    conn.close()

    print()
    print(f"Done! Processed {len(processed_years)} files, {total_rows} total rows")
    print(f"Years in database: {sorted(all_years)}")


if __name__ == '__main__':
    main()

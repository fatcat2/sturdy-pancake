from pymongo.errors import BulkWriteError
from  pymongo import MongoClient
from openpyxl import Workbook
from openpyxl import load_workbook
import os
client = MongoClient('mongodb+srv://exponent_cand:'+os.environ.get('password')+'@cluster0-ufki0.mongodb.net/test')

db = client["test"]
money = db.seventeen
wb = load_workbook(filename='2017.xlsx')
print(wb.sheetnames)

sheet = wb.active

print(sheet['B2'].value)

data_list = []

for row in sheet.iter_rows(min_row=2, min_col=1, max_row=16695, max_col=11):
    a = []
    for cell in row:
        a.append(cell.value)
    print type(a)
    post_data = {
            'last_name': a[1],
            'first_name': a[2],
            'middle_name': a[4],
            'dept': a[4],
            'city': a[6],
            'employee_group': a[9],
            'compensation': a[10]
            }
    data_list.append(post_data)
    #insert the entire fucking list yooooo
try:
    money.insert_many(data_list)
except BulkWriteError as bwe:
    print(bwe.details)
